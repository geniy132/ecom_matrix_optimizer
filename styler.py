import json                                                           # Модуль для чтения файлов конфигурации JSON
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side # Модуль корпоративного дизайна Excel
from openpyxl.utils import get_column_letter                          # Служебный инструмент конвертации индексов


def apply_corporate_style(worksheet, df):
    """Применяет чистый openpyxl дизайн с вертикальным центрированием и подсветкой категорий."""
    # 1. ЗАГРУЗКА БИЗНЕС-ПАЛИТРЫ ИЗ ВНЕШНЕГО ФАЙЛА КОНФИГУРАЦИИ
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            cfg = json.load(f)                                        # Загружаем JSON-конфиг в словарь
    except Exception:
        cfg = {}                                                      # Фолбэк на случай повреждения диска

    # Извлекаем HEX-коды пастельных заливок (убираем знак '#' для openpyxl, если он есть)
    palette_raw = cfg.get("CSS_PALETTE", {})
    def get_clean_hex(key, default):
        raw_val = palette_raw.get(key, default)
        clean_hex = "".join(c for c in raw_val if c.isalnum()).upper()
        return clean_hex[-6:] if len(clean_hex) >= 6 else default

    # Формируем объекты PatternFill для openpyxl на основе вашего config.json
    fill_driver = PatternFill(fill_type='solid', fgColor=get_clean_hex("DRIVER", "E2EFDA"))   # Пастельный зеленый
    fill_star = PatternFill(fill_type='solid', fgColor=get_clean_hex("STAR", "FFF2CC"))       # Пастельный желтый
    fill_turnover = PatternFill(fill_type='solid', fgColor=get_clean_hex("TURNOVER", "FFF2CC")) # Пастельный желтый
    fill_waste = PatternFill(fill_type='solid', fgColor=get_clean_hex("WASTE", "FCE4D6"))     # Пастельный красный
    fill_white = PatternFill(fill_type='solid', fgColor="FFFFFF")                             # Обычный белый

    # Стилизация шрифтов, шапки и контуров сетки
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF') # Белый жирный шрифт заголовков
    font_body = Font(name='Calibri', size=11, bold=False, color='000000') # Обычный шрифт для тела матрицы
    fill_header = PatternFill(fill_type='solid', fgColor='1F4E78')    # Темно-синяя заливка шапки отчета ТЗ
    
    thin_border_side = Side(border_style='thin', color='D9D9D9')      # Настройка тонких серых линий сетки Excel
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Двухмерное выравнивание с вертикальным центрированием строк
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True) # Шапка строго по центру
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True) # Наименования влево
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True) # Числа по центру

    # 2. ФОРМАТИРОВАНИЕ ЗАГОЛОВКОВ ШАПКИ EXCEL ДОКУМЕНТА (ЖЕЛЕЗОБЕТОННЫЙ ЦИКЛ ITER_ROWS)
    worksheet.row_dimensions[1].height = 28                           # Задаем высоту строго для первой строки
    for row in worksheet.iter_rows(min_row=1, max_row=1):             # Берем строго первую строку таблицы
        for cell in row:                                              # Перебираем каждую ячейку в этой строке
            cell.font = font_header                                   # Накладываем жирный белый шрифт Calibri
            cell.fill = fill_header                                   # Заливаем ячейку темно-синим e-com цветом
            cell.alignment = align_header                             # Центрируем текст во всех плоскостях
            cell.border = border_cell                                 # Отрисовываем тонкую серую рамку сетки

    # Находим индекс колонки с рекомендациями, чтобы определять цвет строки
    col_headers = [str(worksheet.cell(row=1, column=c).value).lower().strip() for c in range(1, worksheet.max_column + 1)]
    rec_col_idx = next((i + 1 for i, h in enumerate(col_headers) if "рекомендация" in h or "вывод по акции" in h), None)

    # 3. ПОСТРОЧНЫЙ ДИЗАЙН, УМНАЯ ПОДСВЕТКА И ЦЕНТРИРОВАНИЕ ТЕЛА ТАБЛИЦЫ
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        # По умолчанию красим строку в белый цвет
        current_row_fill = fill_white

        # Если колонка рекомендаций найдена, динамически определяем цвет строки на основе инсайтов
        if rec_col_idx:
            rec_val = str(worksheet.cell(row=row_idx, column=rec_col_idx).value).lower()
            if "драйвер" in rec_val:
                current_row_fill = fill_driver
            elif "звезда" in rec_val:
                current_row_fill = fill_star
            elif "оборотный" in rec_val:
                current_row_fill = fill_turnover
            elif "балласт" in rec_val or "провалена" in rec_val:
                current_row_fill = fill_waste

        worksheet.row_dimensions[row_idx].height = None              # Сбрасываем высоту для автоподбора текста

        for cell in row:
            cell.font = font_body                                     # Применяем стандартный читаемый шрифт Calibri
            cell.fill = current_row_fill                              # Накладываем вычисленный интеллектуальный цвет
            cell.border = border_cell                                 # Применяем аккуратную серую рамку сетки

            # Умное адаптивное распределение горизонтального выравнивания
            if cell.column == 1:
                cell.alignment = align_left                           # Текст влево, центрирование по вертикали
            else:
                cell.alignment = align_center                         # Все числовые метрики — идеально по центру

            # Динамическое наложение числовых бухгалтерских форматов ячеек
            col_name = str(worksheet.cell(row=1, column=cell.column).value).lower()
            if "выруч" in col_name or "прибыль" in col_name or "цена" in col_name:
                cell.number_format = '#,##0'                          # Бухгалтерский разделитель тысяч рублей
            elif "шт" in col_name or "продано" in col_name or "количество" in col_name:
                cell.number_format = '#,##0'                          # Целочисленный формат штук
            elif "%" in col_name or "маржа" in col_name:
                cell.number_format = '0.0'                            # Процентный формат с 1 знаком после точки

    # 4. ДИНАМИЧЕСКИЙ АВТОПОДБОР ШИРИНЫ КОЛОНОК ПО ДЛИНЕ КОНТЕНТА
    for col_idx in range(1, worksheet.max_column + 1):                # Поочередно перебираем индексы всех столбцов листа
        max_len = 0                                                   # Инициализируем стартовую длину строки
        col_letter = get_column_letter(col_idx)                       # Извлекаем буквенное имя столбца Excel (A, B, C...)

        for row_idx in range(1, worksheet.max_row + 1):               # Вертикально сканируем каждую строчку этого столбца
            cell_val = worksheet.cell(row=row_idx, column=col_idx).value # Считываем сырое значение ячейки
            val_str = str(cell_val or '')                             # Принудительно переводим содержимое в текст

            if '\n' in val_str:                                       # Если внутри ячейки есть переносы строк —
                lines = val_str.split('\n')                           # расщепляем текст на отдельные строчки
                max_len = max(max_len, max(len(ln) for ln in lines))  # ищем пиковую длину самой длинной строки контента
            else:
                max_len = max(max_len, len(val_str))                  # Иначе берем стандартный размер контента ячейки
 
        # Задаем ширину столбца с безопасным запасом, ограничивая пик на 45 символах
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)
