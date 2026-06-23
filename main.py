import os                                                             # Модуль файловой системы ОС
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # Инструменты стилизации Excel
import pandas as pd                                                   # Аналитическая библиотека таблиц

from core import (
    calculate_demand_forecast,
    calculate_marginal_profit,
    generate_business_recommendations,
)                                                                     # Импорт расчетных формул ядра
from utils import (
    clean_numeric_string,
    extract_date_columns,
    find_table_header,
)                                                                     # Импорт инструментов парсинга и ETL


def run_pipeline(file_path):
    print("\n🚀 Запуск модульного аналитического конвейера...")

    header_idx = find_table_header(file_path)                         # Динамический поиск шапки
    df = pd.read_excel(file_path, header=header_idx)                  # Чтение исходного Excel
    df.columns = [str(c).strip() for c in df.columns]                 # Очистка имен колонок от пробелов

    # Умный поиск основных бизнес-столбцов по синонимам
    name_col = next((c for c in df.columns if any(x in c.lower() for x in ["полное название", "полное наименование"])), None)
    if not name_col:
        name_col = next((c for c in df.columns if any(x in c.lower() for x in ["название", "товар", "наименование"])), None)
    sku_col = next((c for c in df.columns if any(x in c.lower() for x in ["sku", "артикул"])), None)
    price_col = next((c for c in df.columns if "цена" in c.lower()), None)
    margin_col = next((c for c in df.columns if "маржа" in c.lower()), None)

    if not all([name_col, sku_col, price_col, margin_col]):            # Валидация структуры файла
        print("❌ Ошибка: Критические столбцы не найдены.")
        return 

    df = df.dropna(subset=[name_col, sku_col]).copy()                 # Удаление NaN-строк с копированием
    df = df[df[sku_col].astype(str).str.contains(r"^\d+$", regex=True)].copy() # Фильтр "Итого" по маске SKU

    # Передаем бизнес-колонки, чтобы исключить их из поиска временных рядов продаж
    business_cols = [name_col, sku_col, price_col, margin_col]
    date_list = extract_date_columns(df, business_cols)

    date_cols = [item[0] for item in date_list]                       # Извлекаем только оригинальные имена колонок
    count_days = len(date_cols)                                       # Длина временного ряда (дни)

    if count_days == 0:                                                # Проверка наличия данных продаж
        print("❌ Ошибка: Временные ряды продаж не обнаружены.")
        return

    _, last_date = date_list[-1]
    print(f"📊 Анализируем временной ряд: {count_days} дней. Последняя дата в файле: {last_date.strftime('%d.%m.%Y')}")

    for d_col in date_cols:                                            # Приведение ячеек дат к числу
        df[d_col] = pd.to_numeric(df[d_col], errors="coerce").fillna(0)

    df["_calc_price"] = df[price_col].apply(clean_numeric_string)     # Очистка цен во временный столбец
    df["_calc_margin_raw"] = df[margin_col].apply(clean_numeric_string) # Очистка маржи во временный столбец
    df["_calc_margin_pct"] = df["_calc_margin_raw"].apply(lambda x: x / 100 if x > 1 else x)

    df, forecast_cols = calculate_demand_forecast(df, date_cols, count_days, last_date)
    df = calculate_marginal_profit(df)
    df = generate_business_recommendations(df)

    output_structure = {
        name_col: "Наименование товара",
        sku_col: "SKU",
        price_col: "Цена розничная, руб",
        margin_col: "Маржа, %",
        "Всего_продано_шт": "Продано за период, шт",
    }

    for f_col, _ in forecast_cols:
        output_structure[f_col] = f_col

    output_structure["Маржинальная прибыль за период, руб"] = "Маржинальная прибыль за период, руб"
    output_structure["Рекомендация по оптимизации"] = "Рекомендация по оптимизации"

    final_df = df[list(output_structure.keys())].rename(columns=output_structure)

    output_dir = os.path.dirname(file_path)
    output_file = os.path.join(output_dir, "Итоговый_анализ_ТЗ.xlsx") if output_dir else "Итоговый_анализ_ТЗ.xlsx"

    # Запись в Excel с использованием openpyxl-движка для последующего дизайна
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, sheet_name='Аналитика матрицы')
        worksheet = writer.sheets['Аналитика матрицы']                 # Доступ к рабочему листу

        # Палитра пастельных цветов для подсветки результатов
        fill_driver = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')   # Мягкий зеленый
        fill_star = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')     # Пастельный желтый
        fill_waste = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')    # Пастельный красный

        # Стилизация шапки (темно-синий фон, белый жирный шрифт)
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

        # Общие стили ячеек данных
        data_font = Font(name='Calibri', size=11, bold=False)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Границы ячеек (сетка таблицы)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        # 1. Стилизация строки заголовков (первая строка таблицы)
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        # Находим индекс колонки с рекомендациями
        rec_col_idx = list(final_df.columns).index("Рекомендация по оптимизации") + 1

        # 2. Построчный обход данных (начиная со 2-й строки) и условная заливка
        for row_idx in range(2, worksheet.max_row + 1):
            rec_value = str(worksheet.cell(row=row_idx, column=rec_col_idx).value).lower()

            # Выбор цвета на основе аналитического вердикта
            current_fill = None
            if "драйвер" in rec_value:
                current_fill = fill_driver
            elif "звезда" in rec_value or "оборотный" in rec_value:
                current_fill = fill_star
            elif "балласт" in rec_value:
                current_fill = fill_waste

            # Применяем шрифты, границы и условную заливку построчно
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if current_fill:
                    cell.fill = current_fill
                cell.font = data_font
                cell.border = thin_border
                # Первую колонку (название) выравниваем влево, остальные — по центру
                cell.alignment = align_left if col_idx == 1 else align_center

        # 3. Автоматическое масштабирование ширины столбцов под длину содержимого
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter                       # Безопасное извлечение буквы столбца
            # Ограничиваем размеры: минимум 12, maximum 45 символов в ширину
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    print(f"🎉 Проект успешно выполнен. Красивый структурированный отчет сохранен в: {output_file}")


if __name__ == "__main__":
    print("=== Утилита автоматизации маржинального e-com анализа ===")
    user_input = input("👉 Перетащите файл Excel в это окно или вставьте путь (ПКМ): ").strip()
    clean_path = user_input.strip("'\"")

    if os.path.exists(clean_path):                                     # Проверка наличия указанного файла
        try:
            run_pipeline(clean_path)                                   # Запуск аналитического пайплайна
        except Exception as e:                                         # Изолирование непредвиденных сбоев
            print(f"❌ Критический сбой при расчете: {e}")
    else:
        print(f"❌ Ошибка: Файл по пути '{clean_path}' не найден.")

    input("\nНажмите Enter для выхода...")                            # Удержание консольного окна открытым
